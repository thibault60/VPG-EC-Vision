import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json
import requests
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="VP SEO Dashboard", page_icon="🗺️", layout="wide")

# ──────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────

CATEGORY_RULES = [
    ("All inclusive",   ["all-inclusive", "tout-compris", "tout-inclus",
                         "vacances-et-sejours-all-inclusive", "bali-tout-compris",
                         "voyage-tout-compris", "voyage-tout-inclus", "derniere-minute-tout-inclus",
                         "week-end-tout-compris", "sejours-tout-compris", "voyage-derniere-minute-tout-inclus"]),
    ("Circuit",         ["circuit-", "autotour", "circuits-", "voyage-organise-"]),
    ("Dernière minute", ["last-minute", "derniere-minute", "dernieres-minutes",
                         "depart-demain", "derniere-chance", "croisieres-sur-le-nil-last-minute"]),
    ("Vol + Hôtel",     ["vol-hotel", "vols-plus-hotels", "vol-plus-hotel"]),
    ("Hôtel",           ["hotel-", "hotels-", "hotel-en-", "hotel-a-", "hotel-au-", "hotel-all-inclusive"]),
    ("Meilleurs",       ["top-", "meilleurs-hotels", "meilleurs-hotels-clubs",
                         "les-meilleurs", "top-5-", "top-9-", "top-10-"]),
    ("Où partir",       ["ou-partir", "Ou-partir", "ou-partir-", "quand-partir"]),
    ("Que faire",       ["que-faire"]),
    ("Séjour",          ["sejour-", "sejours-", "offres-sejours", "sejour-a-", "sejour-au-",
                         "sejour-en-", "sejours-en-", "sejours-aux-", "sejours-de-luxe", "sejour-et-voyage-"]),
    ("Ski",             ["ski-"]),
    ("Ventes privées",  ["vente-privee", "ventes-privees", "promo-", "bon-plan", "bons-plans", "bons-plan",
                         "offres-black-friday", "sejour-et-voyage-pas-cher", "vente-privee-pour", "promo-vacances"]),
    ("Vacances",        ["vacances-", "vacance-", "vacances-famille", "vacances-pas-cheres"]),
    ("Week-end",        ["week-end-", "week-ends-", "week-end-a-", "week-end-au-", "week-end-tout-compris"]),
    ("Voyage",          ["voyage-", "voyages-", "voyage-au-", "voyage-en-", "voyage-aux-", "voyage-dans-"]),
    ("Combiné",         ["combine-", "combined-"]),
    ("Croisière",       ["croisiere", "nil-last-minute"]),
    ("Home / Login",    ["login", "login/index"]),
    ("Autre",           ["voyage-pirates", "offres-black-friday"]),
]

MONTH_LABELS = {
    1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
    5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
    9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
}

MANUAL_URL_MAPPING = {
    21271: "https://www.voyage-prive.com/login/index",
}

# ──────────────────────────────────────────
# FONCTIONS UTILITAIRES
# ──────────────────────────────────────────

def normalize_url(url):
    """Lowercase + strip trailing slash — pour le matching GSC ↔ Carambola."""
    if pd.isna(url) or not isinstance(url, str):
        return None
    return url.lower().rstrip("/").strip()

def categorize_url_rules(url):
    if not url or not isinstance(url, str):
        return "Autre"
    if "login" in url:
        return "Home / Login"
    slug = url.lower().split("/offres/")[-1] if "/offres/" in url else url.lower()
    for cat, patterns in CATEGORY_RULES:
        for p in patterns:
            if slug.startswith(p.lower()) or p.lower() in slug:
                return cat
    return None

def extract_destination_rules(slug):
    slug = slug.lower()
    destinations = {
        "France":       ["france", "corse", "paris", "bordeaux", "lyon", "guyane", "martinique", "guadeloupe", "antilles", "reunion"],
        "Maroc":        ["maroc", "marrakech", "agadir", "casablanca"],
        "Espagne":      ["espagne", "barcelone", "madrid", "seville", "majorque", "mallorca", "minorque", "ibiza", "canaries", "lanzarote", "fuerteventura", "tenerife", "andalousie"],
        "Grèce":        ["grece", "crete", "santorin", "mykonos", "rhodes", "corfou", "iles-grecques"],
        "Italie":       ["italie", "rome", "venise", "sardaigne", "sicile", "pouilles", "cinque-terre", "capri", "amalfi", "florence"],
        "Egypte":       ["egypte", "hurghada", "charm-el-cheikh", "louxor"],
        "Maldives":     ["maldives"],
        "Thaïlande":    ["thailande", "bangkok", "phuket", "koh-samui"],
        "Dubai":        ["dubai", "emirats-arabes"],
        "Turquie":      ["turquie", "istanbul", "cappadoce", "antalya", "bodrum"],
        "Tunisie":      ["tunisie", "djerba", "hammamet"],
        "Portugal":     ["portugal", "lisbonne", "porto", "madere", "acores"],
        "Mexique":      ["mexique", "cancun", "playa-del-carmen", "riviera-maya"],
        "Bali":         ["bali"],
        "Japon":        ["japon", "tokyo", "kyoto", "osaka"],
        "Kenya":        ["kenya"],
        "Zanzibar":     ["zanzibar"],
        "Seychelles":   ["seychelles"],
        "Île Maurice":  ["ile-maurice", "mauric"],
        "Sri Lanka":    ["sri-lanka"],
        "Jordanie":     ["jordanie", "petra"],
        "Croatie":      ["croatie", "dubrovnik"],
        "Islande":      ["islande", "reykjavik"],
        "Laponie":      ["laponie", "abisko"],
        "Norvège":      ["norvege"],
        "Suisse":       ["suisse", "geneve"],
        "Belgique":     ["belgique", "bruxelles"],
        "Albanie":      ["albanie"],
        "Cuba":         ["cuba", "la-havane"],
        "Bahamas":      ["bahamas"],
        "Pérou":        ["perou", "lima", "machu-picchu"],
        "Malaisie":     ["malaisie", "kuala-lumpur"],
        "Namibie":      ["namibie"],
        "Tanzanie":     ["tanzanie"],
        "Afrique du Sud": ["afrique-du-sud", "cape-town"],
        "Cap-Vert":     ["cap-vert", "sal"],
        "Hawaii":       ["hawaii"],
        "Punta Cana":   ["punta-cana"],
        "Budapest":     ["budapest"],
        "Prague":       ["prague"],
        "Amsterdam":    ["amsterdam"],
        "Londres":      ["londres", "london"],
        "New York":     ["new-york"],
        "Finlande":     ["finlande"],
        "Europe":       ["europe"],
        "Afrique":      ["afrique"],
        "USA":          ["usa", "etats-unis", "amerique"],
    }
    for dest, keywords in destinations.items():
        for kw in keywords:
            if kw in slug:
                return dest
    return None

def rewrite_url(carambola_url):
    if pd.isna(carambola_url):
        return None
    prefix = "http://fr-carambola.bovpg.net//campaign/microsite/"
    slug = str(carambola_url).replace(prefix, "")
    if slug == "index":
        return f"https://www.voyage-prive.com/login/{slug}"
    return f"https://www.voyage-prive.com/offres/{slug}"

def url_label(url):
    if pd.isna(url):
        return None
    return str(url).replace("https://www.voyage-prive.com/", "")

def parse_float(val):
    if pd.isna(val):
        return None
    try:
        return float(str(val).replace(",", "."))
    except ValueError:
        return None

def slug_from_url(url):
    if pd.isna(url):
        return None
    return str(url).replace("https://www.voyage-prive.com/offres/", "")\
                   .replace("https://www.voyage-prive.com/login/", "login/")

# ──────────────────────────────────────────
# CATÉGORISATION GPT
# ──────────────────────────────────────────

@st.cache_data(ttl=86400, show_spinner=False)
def categorize_urls_gpt(urls_tuple, openai_key):
    urls = list(urls_tuple)
    result = {}
    to_gpt = []
    for url in urls:
        cat  = categorize_url_rules(url)
        slug = url.lower().split("/offres/")[-1] if "/offres/" in url else url.lower().split("voyage-prive.com/")[-1]
        dest = extract_destination_rules(slug)
        if cat is not None:
            result[url] = {"type": cat, "destination": dest}
        else:
            to_gpt.append(url)

    _key = openai_key
    if not _key:
        try:
            _key = st.secrets["openai"]["api_key"]
        except Exception:
            pass

    if not to_gpt or not _key:
        for url in to_gpt:
            slug = url.lower().split("/offres/")[-1] if "/offres/" in url else ""
            result[url] = {"type": "Voyage", "destination": extract_destination_rules(slug)}
        return result

    for i in range(0, len(to_gpt), 30):
        batch = to_gpt[i:i+30]
        prompt = (
            "Tu es un expert SEO voyage. Pour chaque URL ci-dessous, retourne un JSON avec:\n"
            '- "type": All inclusive|Circuit|Dernière minute|Hôtel|Meilleurs|Où partir|Quand partir|'
            "Séjour|Ski|Vacances|Ventes privées|Vol + Hôtel|Voyage|Week-end|Que faire|Combiné|Croisière|Autre\n"
            '- "destination": pays/région principale ou null\n\nURLs:\n'
            + "\n".join(batch)
            + '\n\nRéponds UNIQUEMENT JSON valide : {"url1":{"type":"...","destination":"..."}}'
        )
        try:
            resp = requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers={"Authorization": f"Bearer {_key}", "Content-Type": "application/json"},
                json={"model": "gpt-4o-mini", "messages": [{"role": "user", "content": prompt}],
                      "temperature": 0, "response_format": {"type": "json_object"}},
                timeout=30
            )
            result.update(json.loads(resp.json()["choices"][0]["message"]["content"].strip()))
        except Exception:
            for url in batch:
                slug = url.lower().split("/offres/")[-1] if "/offres/" in url else ""
                result[url] = {"type": "Voyage", "destination": extract_destination_rules(slug)}
    return result

# ──────────────────────────────────────────
# CHARGEMENT DES DONNÉES
# ──────────────────────────────────────────

@st.cache_data
def load_data(tableau_file, carambola_file):
    df_cara = pd.read_excel(carambola_file)
    df_cara = df_cara[df_cara["Campaign id"] != 0].copy()
    df_cara["vp_url"] = df_cara["Campaign URL"].apply(rewrite_url)

    inscrit_col = next((c for c in ["Inscrits", "inscrits", "New members", "Leads", "leads"] if c in df_cara.columns), None)
    cols = ["Campaign id", "Campaign name", "vp_url"] + ([inscrit_col] if inscrit_col else [])
    df_cara = df_cara[cols].rename(columns={"Campaign id": "campaign_id", "Campaign name": "campaign_name"})
    if inscrit_col:
        df_cara = df_cara.rename(columns={inscrit_col: "inscrits_cara"})

    df_tab = pd.read_csv(tableau_file, sep=";", engine="python")
    df_tab.columns = df_tab.columns.str.strip()

    if "Campaign Id (h1)" not in df_tab.columns:
        raise ValueError(
            "MAUVAIS_FICHIER : colonne 'Campaign Id (h1)' manquante. "
            f"Colonnes trouvées : {list(df_tab.columns)}"
        )

    df_tab = df_tab.rename(columns={
        "Campaign Id (h1)": "campaign_id",
        "Measure Names":    "metric",
        "Measure Values":   "value",
        "Date granularity": "month",
        "Registration Year": "year"
    })
    df_tab = df_tab[df_tab["metric"].isin(["TTV", "Bookings", "TTV per lead Net", "Inscrits"])].copy()
    df_tab = df_tab[df_tab["campaign_id"].notna()].copy()
    df_tab["campaign_id"] = df_tab["campaign_id"].astype(int)
    df_tab["value"]       = df_tab["value"].apply(parse_float)
    df_tab = df_tab[(df_tab["year"] == 2025) & (df_tab["month"].between(1, 12))].copy()
    df_tab["month_label"] = df_tab["month"].map(MONTH_LABELS)

    df = df_tab.merge(df_cara, on="campaign_id", how="left")

    for cid, url in MANUAL_URL_MAPPING.items():
        df.loc[(df["campaign_id"] == cid) & df["vp_url"].isna(), "vp_url"] = url

    df["slug"]       = df["vp_url"].apply(slug_from_url)
    df["url_label"]  = df["vp_url"].apply(url_label)
    df["url_label"]  = df["url_label"].fillna("ID:" + df["campaign_id"].astype(str))
    df["vp_url_norm"]= df["vp_url"].apply(normalize_url)
    return df

@st.cache_data
def load_gsc(gsc_file):
    df = pd.read_csv(gsc_file)
    df.columns = df.columns.str.strip()
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if "landing" in cl or cl in ("url", "page"):
            col_map[col] = "url"
        elif "date" in cl or "month" in cl:
            col_map[col] = "month"
        elif "click" in cl:
            col_map[col] = "clicks"
    df = df.rename(columns=col_map)
    if "url" not in df.columns or "clicks" not in df.columns:
        st.error("GSC : colonnes 'URL/Landing Page' et 'Url Clicks' requises.")
        return pd.DataFrame()
    df["url_norm"] = df["url"].apply(normalize_url)
    df["clicks"]   = pd.to_numeric(df["clicks"], errors="coerce").fillna(0)
    if "month" in df.columns:
        df["month"] = pd.to_numeric(df["month"], errors="coerce")
    return df

# ──────────────────────────────────────────
# HEADER
# ──────────────────────────────────────────

_VP_LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/wAARCAHCAcIDASIAAhEBAxEB/8QAHQABAAICAwEBAAAAAAAAAAAAAAcIAQYEBQkCA//EAFQQAAIBAwIDBAQJCAUHCgcAAAABAgMEBQYRBxIhCDFBURMiYXEUFhgyQlKBkdIVI1ZigpSh0RczcpOVCXWSorG08CQlJjc4VMHT4fFDU1Vzg7PC/8QAFgEBAQEAAAAAAAAAAAAAAAAAAAEC/8QAFhEBAQEAAAAAAAAAAAAAAAAAAAER/9oADAMBAAIRAxEAPwCsIANNAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAcnE2byGTtbGNxbWzuK0aSrXNVU6UHJpc0pPpGK36t9wHGBN1j2YOJF/Z07ywutN3dtVW9OtRyKnCa81JRaa9xyPkp8VfLB/vz/CREEgnb5KfFXywf78/wj5KfFX6uD/fn+EqoJBO3yU+Kv1cH+/P8I+SnxV+rg/35/hAgkE7fJT4q/Vwf78/wj5KfFX6uD/AH5/hCIJBO3yU+Kv1cH+/P8ACPkp8Vfq4P8Afn+ECCQTt8lPir9XB/vz/CcPJ9mDi1Z0XUo4zHX0lv6lvfwT/1+Xf3AQqDudV6U1LpS9VnqTB3+LrN7RVzRlCM/wCzLbaX2HTMADBkAAAAAAAAAAAAAAAAAAAAMf7fA7LT2BzeoshHH4HG3uUu5LdUrWhKpJLzaS6R9r2QHXAmPEdmfi7f0lVqYK1sIyW6V1fUlL7ouT+/Y7P5KfFX6uD/AH1/hCIJBO3yU+Kv1cH+/P8ACPkp8Vfq4P9+f4QIJBO3yU+Kv1cH+/P8I+SnxV8sH+/P8IEEgnb5KfFXywf78/wnzPsrcUoRcpvAxjFNt/Dn0X2xAgsHfa40vc6SzLxV3k8Rka0afPUljbtXFOm92uSUl3SXK917Tof4gAAAAAAAAAAAAAAAAAAAMeHvWxkAbbw64j6x0DkFd6ZzVe2hKSdW1nLnt63snB9N9t1zd279+/bX7wOt/k2P2sru4sbyjeTWsqtJc9nB6a1Lkn+qy9/mSwAPTKx0tp3TMb3JYqxtbqWFr5GrCnzXlpbKdRp8si2+bZuOyaz8VuImvNXaN0FmNE4XFaLrRp/DbuizSpShyLb0VN8vK3u5Pdt9G1seo/Z04k5bI2F1aYHSGmrW2pqFOhia9VwjyNqPzzqNyko7bJtJHQY/2PtTY7M4bJT660bdUbCtCdawraYqUpXaT3cXzRa3fh+ewEUYPXPEfR2Vx73TGpM3j6Np6aNKhXnGNCK26ekgly+WxyOI9qXiJT0bW0ra3OKucjVq86vpKNOVOEb+GzknuvdMDoaezJoazeL6a+z9pF7e+ULi/m/aovlt7PQjjVv+X9nnTFncQp1tZ5yvXgt+e3owpRb+xpzfuAmLtN8KFxG0X+UMNb8unMrNuvBLpb3HXnS+yW2/wBOT+kQBpHXfBPIaC1TU09qLNYylXp06c1Wpzo2aqRmlzJc0FzRe+3X3M2Lsjaz+MGv6umctGPxH1E5W1x6LZRq7pRb8t/Wj7mQl2htMYzVFribHO21ld47I6ey1C4pV4KcWpVnKEndPb5dRe/i+5bICtnH/WuqsfxJ1Ho/S2jbfU8cVa0cjc3FaVC2jO5iqb3qLbniqa35t9+Vt+7cjEY7I4XO3OH1Nk8lZ2+RuIMxqrGV6dGrJVJqVF80asee55ObblbSXl0JCg7Pva1tLR6v1dlt7qjSgpXkrSMpwT9ZwSim9t9lvtuBDHai4pyv8ARWW0fi7i5sLHI3kridtTqOMKnJKK2fLv82LWzXXbdJM0kY3I5bXN3hq1ximrkqVvCMrZO7hS57aTg3J7NN9enQZgA07BZVW21Dk7XKaYtM9b29OjdWlxhpzjVg1ySlJKUXGUW076vZe8tjhOKWv8AiVgNW5fSmN0hjMXp7VeYtbd06kKtGMK1Srb+smvXr0acVzS3cW1u0S7w14lcO+AOgsXo/QuL+Na3fLXqwpuKqXNVpKVWpJLdttLbulFJJIDs+0jrXXOvP+F9K3i0fp6zp0I/CbFp1ZVqkZSa2nyqKXLttyrc7Lh5petp7V+lMdj6t7fYzPWuajZ3FCNVq2nGpGlJVYyT5ZRUW27R3T9qZrHYX09b6e1dQ1jqvIyrYu2u5XuMjVSlWulJNuVTZ9N/WW6fM007d2wr8PdLan0NkMNR1pktU40tVXj1m7uFZUZz5XKCSahFeTUfVXXdgWehtcXcvZQ4Z3mV4N6GpTy9z8YvquW6qudVaknvu5OXN16r+B2ma1HjNKYO8z2cuFbY6yp+kq1OVy5Yq662i23svBIhfW3HmvofTXELWvD/SWktP1c1kKGOuJ1G6d2oqnGMpW9t0lyzWzfPN7eSX0sDDhrfU0cJLVGMnRlPBq/VJQklNKrGo1Hd7Llkl9YFS+JWtdScQ+MWZ1Tk3UpXFxPaFOLdO1pQ2UacF4bJfW93uyPiQ1ZLT9GdTe6jhaVoqc6mJhW3nKEuV7uS2XNuk9n16HOp1amRuaVrb1ZVq1WUYQpwW8pSb2SXm20egbgTh9T6M4VpajqadqZ6rWvVSpuo6c6VKdSVOi2o7pOOznzLfm5e7cDJpzHanxXBvTukqGqb/F6kr3n5JxN9j4Uk5VOeDq+rVl6uze/JPqlt0PjTmpOItW91TpzTul41tOarRp3cLieQp009lSqSnCSnJ8u+7g3FNbLdbFLpYSLMx1TeyuZWi1CKqVd7qzlUUOVe2cm9vIlnDVNI5KvhPa3yFbTspXkMfP0lBVFVl6N1N5c7iuXZ7dI7Lu8wJ7j2RcBc6HpUb7MasxtW8ptxjOvbOjCh6tKTeTjB1Obl9LfZdXtoQNrHV9bXGq8lqKrRjSle13NQjLmUN9kopu27Sit9+7ck3StLE3uC0zW1/cZ2OlLijV/I1laRqPHQrScJVHCUFJtLeMZc+22/iRvhcvksFkI5LCZCtY3sU4RuLaqoyUZJqS36dU2n7fBgZGSPuDuqtQar0FZ5nU+mZabyanLFW0q0arVBSfLz+jbSU+nmS3XVbp7PbqAAAAAAAAAAAAAAAAAAAAPyx+w7NMZ7JafytplsRdVLW9s6sa1tXpPaSqRe6ku7u260+9Hdn0A7bFdoDiTk7Wra1NX3UaU6UqMotQXSS2fXl8V1NAlCEasK7pQdaEXGM+VcyT22Se23uXuM+AA9K/s86VvNUcOJR1Lk8hkq+Jr+mhC4qyn8GlvFuNJP5seaM3tvvs9t1seHXBLI8S8JqW7sb+wtPyCqlGpCVKU5OtFqXP6q3XotuvMuZp3Y21ljZqWmNf5+0sq0ozpVVTpupTlF7pxcmt1ttuuu5aLSeq8dq/Bt3FhKtbVLWrK1urG5SjVtrqn/S0p7d6e/Rv5SbXRsDRNDdi3GYnXmRzeB1VjbfDXsY+mwNajN+h5VFNpqUoveLcuZf2Wty7mjNG4vROiMRpTES5oWVHkdRNKU6r6znJLq5Pkn9yXTc5mARAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAH//Z"

st.markdown(
    '<div style="display:flex;align-items:center;gap:14px;margin-bottom:8px">'
    '<img src="data:image/png;base64,' + _VP_LOGO_B64 + '" style="height:52px;border-radius:6px">'
    '<span style="font-size:1.8rem;font-weight:700">Voyage Privé · Dashboard SEO Performance</span>'
    '</div>', unsafe_allow_html=True
)
st.markdown("Top pages par mois · TTV, Bookings & Funnel de conversion")

# ──────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────

with st.sidebar:
    st.header("📂 Fichiers de données")
    tableau_file   = st.file_uploader("Export Tableau (CSV)",                         type=["csv"])
    carambola_file = st.file_uploader("Export Carambola (XLSX)",                      type=["xlsx"])
    gsc_file       = st.file_uploader("Export GSC — clics par page/mois (CSV)",       type=["csv"])
    st.divider()
    st.caption("Les fichiers ne sont pas stockés — traitement 100 % local.")

if not tableau_file or not carambola_file:
    st.info("👈 Charge les deux fichiers Tableau + Carambola dans la sidebar pour commencer.")
    st.stop()

df     = load_data(tableau_file, carambola_file)
df_gsc = load_gsc(gsc_file) if gsc_file else pd.DataFrame()

with st.sidebar:
    st.header("🔧 Filtres")
    months_available = sorted(df["month"].unique())
    st.caption("Mois")
    selected_months = [m for m in months_available if st.checkbox(MONTH_LABELS[m], value=True, key=f"mcb_{m}")]
    if not selected_months:
        selected_months = months_available
    top_n = st.slider("Nombre de pages", 5, 500, 50, 5)

    st.divider()
    st.header("🤖 Catégorisation GPT")
    _has_secret = bool(st.secrets.get("openai", {}).get("api_key", ""))
    openai_key  = st.secrets["openai"]["api_key"] if _has_secret else st.text_input("Clé API OpenAI", type="password")
    if _has_secret:
        st.success("✅ Clé OpenAI configurée")
    if st.button("🤖 Lancer la catégorisation IA", use_container_width=True):
        st.session_state["run_cat"] = True
    if st.button("🗑️ Réinitialiser le cache", use_container_width=True):
        st.session_state.pop("run_cat", None)
        st.session_state.pop("cats_cache", None)
        st.cache_data.clear()
    enable_cat = st.session_state.get("run_cat", False)

    st.divider()
    st.header("🗂️ Filtres catégories")

# ──────────────────────────────────────────
# CATÉGORISATION
# ──────────────────────────────────────────

df_filtered = df[df["month"].isin(selected_months)].copy()
ordered_month_labels = [MONTH_LABELS[m] for m in sorted(selected_months)]

all_urls_full = df["vp_url"].dropna().unique().tolist()

if enable_cat:
    if "cats_cache" not in st.session_state:
        with st.spinner("Catégorisation des URLs en cours…"):
            st.session_state["cats_cache"] = categorize_urls_gpt(tuple(all_urls_full), openai_key or "")
    cats = st.session_state["cats_cache"]
    df_filtered["type_page"]   = df_filtered["vp_url"].map(lambda u: cats.get(u, {}).get("type", "Autre") if isinstance(u, str) else "Autre")
    df_filtered["destination"] = df_filtered["vp_url"].map(lambda u: cats.get(u, {}).get("destination") if isinstance(u, str) else None)
else:
    df_filtered["type_page"]   = df_filtered["vp_url"].apply(lambda u: categorize_url_rules(u) or "Voyage")
    df_filtered["destination"] = df_filtered["vp_url"].apply(
        lambda u: extract_destination_rules(u.lower().split("/offres/")[-1] if isinstance(u, str) and "/offres/" in u else "") if isinstance(u, str) else None
    )

all_types = sorted(df_filtered["type_page"].dropna().unique().tolist())
all_dests = sorted(df_filtered["destination"].dropna().unique().tolist())

with st.sidebar:
    filter_type = st.multiselect("Schéma de recherche", options=all_types or sorted({c for c, _ in CATEGORY_RULES}), key="ft")
    filter_dest = st.multiselect("Destination",          options=all_dests, key="fd")

if filter_type:
    df_filtered = df_filtered[df_filtered["type_page"].isin(filter_type)]
if filter_dest:
    df_filtered = df_filtered[df_filtered["destination"].isin(filter_dest)]

# ── KPIs globaux ──
c1, c2 = st.columns(2)
c1.metric("💶 TTV Total", f"{df_filtered[df_filtered['metric']=='TTV']['value'].sum():,.0f} €")
c2.metric("🛎️ Bookings Total", f"{int(df_filtered[df_filtered['metric']=='Bookings']['value'].sum()):,}")
st.divider()

# ──────────────────────────────────────────
# AGRÉGATIONS COMMUNES
# ──────────────────────────────────────────

def get_agg(df_in):
    return df_in.groupby(
        ["month", "month_label", "campaign_id", "campaign_name", "vp_url", "url_label", "slug"],
        dropna=False
    )["value"].sum().reset_index()

df_ttv_agg = get_agg(df_filtered[df_filtered["metric"] == "TTV"])
df_bkg_agg = get_agg(df_filtered[df_filtered["metric"] == "Bookings"])
df_tpl_agg = get_agg(df_filtered[df_filtered["metric"] == "TTV per lead Net"])

# Inscrits : depuis Tableau ou Carambola
if "Inscrits" in df_filtered["metric"].values:
    df_ins_agg = get_agg(df_filtered[df_filtered["metric"] == "Inscrits"])
elif "inscrits_cara" in df_filtered.columns:
    _src = df_filtered[["month","month_label","campaign_id","campaign_name","vp_url","url_label","slug","inscrits_cara"]].drop_duplicates()
    _src = _src.rename(columns={"inscrits_cara": "value"})
    _src["value"] = pd.to_numeric(_src["value"], errors="coerce").fillna(0)
    df_ins_agg = _src.groupby(["month","month_label","campaign_id","campaign_name","vp_url","url_label","slug"], dropna=False)["value"].sum().reset_index()
else:
    df_ins_agg = pd.DataFrame()

_pinned_ids = list(MANUAL_URL_MAPPING.keys())

_ttv_all = df_ttv_agg.groupby(["campaign_id","campaign_name","vp_url","url_label"])["value"].sum().reset_index().rename(columns={"value":"TTV"})
_bkg_all = df_bkg_agg.groupby(["campaign_id","campaign_name","vp_url","url_label"])["value"].sum().reset_index().rename(columns={"value":"Bookings"})
_all     = _ttv_all.merge(_bkg_all, on=["campaign_id","campaign_name","vp_url","url_label"], how="outer").fillna(0)

_active  = _all[(_all["TTV"] > 0) | (_all["Bookings"] > 0)]
_pinned  = _all[_all["campaign_id"].isin(_pinned_ids)]
_clean   = pd.concat([_active, _pinned]).drop_duplicates(subset=["campaign_id"])

_rest          = _clean[~_clean["campaign_id"].isin(_pinned_ids)].sort_values("TTV", ascending=False).head(top_n)
df_ttv_global  = pd.concat([_clean[_clean["campaign_id"].isin(_pinned_ids)], _rest]).drop_duplicates(subset=["campaign_id"])
top_ids        = df_ttv_global["campaign_id"].tolist()
df_merged_global = df_ttv_global.sort_values("TTV", ascending=True)

palette   = px.colors.qualitative.Plotly + px.colors.qualitative.Dark24 + px.colors.qualitative.Light24
color_map = {
    row["url_label"]: palette[i % len(palette)]
    for i, (_, row) in enumerate(
        df_ttv_agg[df_ttv_agg["campaign_id"].isin(top_ids)]
        .groupby(["campaign_id","url_label"])["value"].sum()
        .reset_index().sort_values("value", ascending=False).iterrows()
    )
}

# ──────────────────────────────────────────
# ONGLETS
# ──────────────────────────────────────────

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Top URLs", "📅 Vue par mois", "🌍 Par destination", "🗂️ Par type de page", "📈 Funnel & Conversion"
])

# ─── TAB 1 : TOP URLs ───────────────────

with tab1:
    st.subheader(f"📊 Top {top_n} URLs · TTV & Bookings")

    df_donut = df_merged_global.sort_values("TTV", ascending=False).reset_index(drop=True)
    df_donut["color"] = [palette[i % len(palette)] for i in range(len(df_donut))]

    col_d, col_t = st.columns(2)
    with col_d:
        fig_d = go.Figure(go.Pie(
            labels=df_donut["url_label"], values=df_donut["TTV"], hole=0.45,
            marker=dict(colors=df_donut["color"]), textinfo="percent", sort=False,
            hovertemplate="<b>%{label}</b><br>TTV : %{value:,.0f} €<br>Part : %{percent}<extra></extra>"
        ))
        fig_d.update_layout(title=f"TTV · Top {top_n} URLs", height=600, showlegend=False,
                            margin=dict(l=10, r=10, t=50, b=10))
        st.plotly_chart(fig_d, use_container_width=True)

    with col_t:
        st.markdown(f"**Bookings · Top {top_n} URLs**")
        df_bd = df_donut.sort_values("Bookings", ascending=False)
        rows  = "".join(
            f"<tr><td><span style='display:inline-block;width:11px;height:11px;border-radius:50%;"
            f"background:{r.color};margin-right:6px'></span>{r.url_label}</td>"
            f"<td style='text-align:right'><b>{int(r.Bookings):,}</b></td>"
            f"<td style='text-align:right;color:#888'>{r.TTV:,.0f} €</td></tr>"
            for r in df_bd.itertuples()
        )
        st.markdown(
            f"<style>.bkt{{width:100%;border-collapse:collapse;font-size:12.5px}}"
            f".bkt th{{padding:6px 8px;border-bottom:2px solid #ddd;color:#555;font-size:12px}}"
            f".bkt td{{padding:4px 8px;border-bottom:1px solid #f0f0f0}}"
            f".bkt tr:hover td{{background:#f7f7f7}}</style>"
            f"<div style='max-height:560px;overflow-y:auto'><table class='bkt'>"
            f"<thead><tr><th>URL</th><th style='text-align:right'>Bookings</th><th style='text-align:right'>TTV (€)</th></tr></thead>"
            f"<tbody>{rows}</tbody></table></div>",
            unsafe_allow_html=True
        )

    st.subheader(f"📈 Évolution mensuelle · Top {top_n} URLs")
    ev_mode = st.radio("Afficher :", ["TTV + Bookings", "TTV uniquement", "Bookings uniquement"],
                       horizontal=True, label_visibility="collapsed", key="ev1")

    fig_ev = make_subplots(specs=[[{"secondary_y": True}]])
    if ev_mode in ("TTV + Bookings", "TTV uniquement"):
        for cid in top_ids:
            lbl = (df_merged_global[df_merged_global["campaign_id"]==cid]["url_label"].values or [f"ID:{cid}"])[0]
            r   = df_ttv_agg[df_ttv_agg["campaign_id"]==cid].sort_values("month")
            fig_ev.add_trace(go.Bar(x=r["month_label"], y=r["value"], name=lbl,
                                    marker=dict(color=color_map.get(lbl,"#888")), legendgroup=lbl,
                                    hovertemplate=f"<b>{lbl}</b><br>TTV : %{{y:,.0f}} €<extra></extra>"),
                             secondary_y=False)
    if ev_mode in ("TTV + Bookings", "Bookings uniquement"):
        for cid in top_ids:
            lbl = (df_merged_global[df_merged_global["campaign_id"]==cid]["url_label"].values or [f"ID:{cid}"])[0]
            r   = df_bkg_agg[df_bkg_agg["campaign_id"]==cid].sort_values("month")
            fig_ev.add_trace(go.Scatter(x=r["month_label"], y=r["value"], name=lbl, mode="lines+markers",
                                        line=dict(color=color_map.get(lbl,"#888"), width=1.8, dash="dot"),
                                        marker=dict(size=5), legendgroup=lbl,
                                        showlegend=(ev_mode=="Bookings uniquement"),
                                        hovertemplate=f"<b>{lbl}</b><br>Bookings : %{{y:,.0f}}<extra></extra>"),
                             secondary_y=True)
    fig_ev.update_layout(barmode="stack", height=700,
                         xaxis=dict(categoryorder="array", categoryarray=ordered_month_labels, tickangle=-30),
                         legend=dict(orientation="h", yanchor="top", y=-0.2, font=dict(size=9)),
                         margin=dict(t=40, b=180, l=60, r=60))
    fig_ev.update_yaxes(title_text="TTV (€)", secondary_y=False)
    fig_ev.update_yaxes(title_text="Bookings", secondary_y=True, showgrid=False)
    st.plotly_chart(fig_ev, use_container_width=True)

    st.subheader("🗓️ Heatmap mensuelle")
    hc1, hc2 = st.columns(2)
    for hc, hmet, hlbl, hcol in [(hc1,"TTV","TTV (€)","Blues"),(hc2,"Bookings","Bookings","Oranges")]:
        with hc:
            src = df_filtered[df_filtered["metric"]==hmet].copy()
            src = src[src["vp_url"].notna()].copy()
            src["_u"] = src["vp_url"].str.replace("https://www.voyage-prive.com/","",regex=False)
            piv = src.groupby(["_u","month"])["value"].sum().reset_index()\
                     .pivot(index="_u", columns="month", values="value").fillna(0)
            piv["_t"] = piv.sum(axis=1)
            piv = piv.sort_values("_t", ascending=False).head(top_n).drop(columns="_t")
            piv.columns = [MONTH_LABELS.get(c, str(c)) for c in piv.columns]
            fig_hm = px.imshow(piv, color_continuous_scale=hcol, aspect="auto",
                               title=f"Heatmap {hlbl} · Top {top_n}", text_auto=".3s",
                               labels=dict(x="Mois", y="URL", color=hlbl))
            fig_hm.update_layout(height=max(400,top_n*30), coloraxis_showscale=False,
                                 xaxis=dict(side="top"), margin=dict(l=10,r=10,t=50,b=10))
            fig_hm.update_traces(textfont_size=9)
            st.plotly_chart(fig_hm, use_container_width=True)

# ─── TAB 2 : VUE PAR MOIS ───────────────

with tab2:
    st.subheader("📅 Vue détaillée par mois")
    sel_m = st.selectbox("Mois", options=sorted(selected_months), format_func=lambda x: MONTH_LABELS[x])
    dm    = df_filtered[df_filtered["month"] == sel_m]
    dm_ttv = dm[dm["metric"]=="TTV"].groupby(["campaign_id","campaign_name","vp_url","url_label"])["value"]\
               .sum().reset_index().rename(columns={"value":"TTV"})\
               .sort_values("TTV", ascending=False).head(top_n)
    dm_bkg = dm[dm["metric"]=="Bookings"].groupby(["campaign_id","vp_url","url_label"])["value"]\
               .sum().reset_index().rename(columns={"value":"Bookings"})
    dm_mg  = dm_ttv.merge(dm_bkg, on=["campaign_id","vp_url","url_label"], how="left").fillna({"Bookings":0})
    dm_mg  = dm_mg.sort_values("TTV", ascending=True)

    fig_m = go.Figure()
    fig_m.add_trace(go.Bar(y=dm_mg["url_label"], x=dm_mg["TTV"],    name="TTV (€)",  orientation="h", marker_color="#1d6fa4", opacity=0.85))
    fig_m.add_trace(go.Bar(y=dm_mg["url_label"], x=dm_mg["Bookings"],name="Bookings", orientation="h", marker_color="#f4a840", opacity=0.85))
    fig_m.update_layout(barmode="group", height=max(400,top_n*45),
                        title=f"Top {top_n} URLs · {MONTH_LABELS[sel_m]}",
                        yaxis=dict(tickfont=dict(size=10)), xaxis=dict(title="Valeur"),
                        legend=dict(orientation="h", yanchor="bottom", y=1.02),
                        margin=dict(l=20,r=40,t=60,b=20))
    st.plotly_chart(fig_m, use_container_width=True)
    st.dataframe(dm_mg[["campaign_id","campaign_name","vp_url","TTV","Bookings"]]
                 .rename(columns={"campaign_id":"ID","campaign_name":"Campagne","vp_url":"URL VP"}),
                 hide_index=True, use_container_width=True)

# ─── TAB 3 : PAR DESTINATION ────────────

with tab3:
    _dc = df_filtered[df_filtered["destination"].notna()].copy() if "destination" in df_filtered.columns else pd.DataFrame()
    if _dc.empty:
        st.info("Active la catégorisation pour voir les données par destination.")
    else:
        _td = _dc[_dc["metric"]=="TTV"].groupby("destination")["value"].sum().reset_index().sort_values("value",ascending=False)
        _bd = _dc[_dc["metric"]=="Bookings"].groupby("destination")["value"].sum().reset_index().sort_values("value",ascending=False)
        _ld = _dc[_dc["metric"]=="TTV per lead Net"].groupby("destination")["value"].mean().reset_index().sort_values("value",ascending=False)
        c1d,c2d,c3d = st.columns(3)
        for col, df_d, title, scale in [
            (c1d,_td,"TTV (€) par destination","Blues"),
            (c2d,_bd,"Bookings par destination","Oranges"),
            (c3d,_ld,"TTV / Lead (€) par destination","Greens"),
        ]:
            with col:
                val_col = df_d.columns[1]
                fig = px.bar(df_d, x=val_col, y="destination", orientation="h", title=title,
                             color=val_col, color_continuous_scale=scale,
                             height=max(400,len(df_d)*28), labels={val_col: title.split(" par")[0]})
                fig.update_layout(yaxis=dict(autorange="reversed"), coloraxis_showscale=False)
                st.plotly_chart(fig, use_container_width=True)

        st.subheader("📈 Évolution mensuelle par destination")
        ev_dm = st.radio("Métrique", ["TTV (€)","Bookings"], horizontal=True, key="evdm")
        _me   = "TTV" if ev_dm == "TTV (€)" else "Bookings"
        _evd  = _dc[_dc["metric"]==_me].groupby(["month","month_label","destination"])["value"].sum().reset_index()
        fig_evd = px.bar(_evd, x="month_label", y="value", color="destination", barmode="stack",
                         category_orders={"month_label": ordered_month_labels}, height=500,
                         labels={"value":ev_dm,"month_label":"Mois","destination":"Destination"})
        fig_evd.update_layout(legend=dict(orientation="h",yanchor="top",y=-0.2,font=dict(size=9)), margin=dict(b=180))
        st.plotly_chart(fig_evd, use_container_width=True)

# ─── TAB 4 : PAR TYPE DE PAGE ───────────

with tab4:
    _tc = df_filtered[df_filtered["type_page"].notna()].copy() if "type_page" in df_filtered.columns else pd.DataFrame()
    if _tc.empty:
        st.info("Active la catégorisation pour voir les données par type de page.")
    else:
        if st.checkbox("Masquer 'Autre'", value=True, key="ha4"):
            _tc = _tc[_tc["type_page"] != "Autre"]
        _tt = _tc[_tc["metric"]=="TTV"].groupby("type_page")["value"].sum().reset_index().sort_values("value",ascending=False)
        _bt = _tc[_tc["metric"]=="Bookings"].groupby("type_page")["value"].sum().reset_index().sort_values("value",ascending=False)
        _lt = _tc[_tc["metric"]=="TTV per lead Net"].groupby("type_page")["value"].mean().reset_index().sort_values("value",ascending=False)
        c1t,c2t,c3t = st.columns(3)
        for col, df_t, title, scale in [
            (c1t,_tt,"TTV (€) par type","Blues"),
            (c2t,_bt,"Bookings par type","Oranges"),
            (c3t,_lt,"TTV / Lead par type","Greens"),
        ]:
            with col:
                val_col = df_t.columns[1]
                fig = px.bar(df_t, x=val_col, y="type_page", orientation="h", title=title,
                             color=val_col, color_continuous_scale=scale,
                             height=max(400,len(df_t)*40), labels={val_col:title.split(" par")[0],"type_page":"Type"})
                fig.update_layout(yaxis=dict(autorange="reversed"), coloraxis_showscale=False)
                st.plotly_chart(fig, use_container_width=True)

        st.subheader("📈 Évolution mensuelle par type de page")
        ev_tm = st.radio("Métrique", ["TTV (€)","Bookings"], horizontal=True, key="evtm")
        _me   = "TTV" if ev_tm == "TTV (€)" else "Bookings"
        _evt  = _tc[_tc["metric"]==_me].groupby(["month","month_label","type_page"])["value"].sum().reset_index()
        fig_evt = px.bar(_evt, x="month_label", y="value", color="type_page", barmode="stack",
                         category_orders={"month_label": ordered_month_labels}, height=500,
                         labels={"value":ev_tm,"month_label":"Mois","type_page":"Type"})
        fig_evt.update_layout(legend=dict(orientation="h",yanchor="top",y=-0.2,font=dict(size=9)), margin=dict(b=180))
        st.plotly_chart(fig_evt, use_container_width=True)

# ─── TAB 5 : FUNNEL & CONVERSION ────────

with tab5:
    st.subheader("📈 Funnel & Conversion — Trafic GSC → Inscrits → Bookings → TTV")

    if df_gsc.empty:
        st.info("👈 Charge le fichier export GSC dans la sidebar pour activer cet onglet.\n\n"
                "**Format attendu :** `Landing Page, Date (Month), Url Clicks`")
    else:
        # Filtrage GSC sur les mois sélectionnés
        gf = df_gsc.copy()
        if "month" in gf.columns and gf["month"].notna().any():
            gf = gf[gf["month"].isin(selected_months)]

        gsc_agg = gf.groupby("url_norm")["clicks"].sum().reset_index()

        # Agréger chaque métrique par URL normalisée
        def norm_agg(agg_df, col):
            if agg_df.empty:
                return pd.DataFrame(columns=["vp_url_norm", col])
            t = agg_df.copy()
            t["vp_url_norm"] = t["vp_url"].apply(normalize_url)
            return t.groupby("vp_url_norm")["value"].sum().reset_index().rename(columns={"value": col})

        ins_n = norm_agg(df_ins_agg, "Inscrits") if not df_ins_agg.empty else pd.DataFrame(columns=["vp_url_norm","Inscrits"])
        bkg_n = norm_agg(df_bkg_agg, "Bookings")
        ttv_n = norm_agg(df_ttv_agg, "TTV")

        url_lbl_df = df_filtered[["vp_url","vp_url_norm","url_label","type_page"]].drop_duplicates("vp_url_norm")

        # Fusion
        fdf = gsc_agg.rename(columns={"url_norm": "vp_url_norm"})
        for mdf in [ins_n, bkg_n, ttv_n]:
            fdf = fdf.merge(mdf, on="vp_url_norm", how="left")
        fdf = fdf.merge(url_lbl_df, on="vp_url_norm", how="left")
        fdf = fdf.fillna({"Inscrits": 0, "Bookings": 0, "TTV": 0})
        fdf["url_label"] = fdf["url_label"].fillna(fdf["vp_url_norm"])

        # Métriques de conversion
        fdf["conv_leads_pct"] = (fdf["Inscrits"]  / fdf["clicks"].replace(0, float("nan")) * 100).round(2)
        fdf["conv_bkg_pct"]   = (fdf["Bookings"]  / fdf["Inscrits"].replace(0, float("nan")) * 100).round(2)
        fdf["aov"]            = (fdf["TTV"]        / fdf["Bookings"].replace(0, float("nan"))).round(0)

        fa = fdf[fdf["clicks"] > 0].copy()  # URLs actives (trafic GSC > 0)
        has_ins = fa["Inscrits"].sum() > 0

        # ── KPIs globaux ──
        tot_clk = int(fa["clicks"].sum())
        tot_ins = int(fa["Inscrits"].sum())
        tot_bkg = int(fa["Bookings"].sum())
        tot_ttv = fa["TTV"].sum()
        avg_cl  = round(tot_ins / tot_clk * 100, 2) if tot_clk > 0 else 0
        avg_cb  = round(tot_bkg / tot_ins * 100, 2) if tot_ins > 0 else 0
        avg_aov = round(tot_ttv / tot_bkg, 0)       if tot_bkg > 0 else 0

        k1,k2,k3,k4,k5,k6 = st.columns(6)
        k1.metric("🖱️ Clics GSC",     f"{tot_clk:,}")
        k2.metric("👤 Inscrits",       f"{tot_ins:,}")
        k3.metric("🛎️ Bookings",       f"{tot_bkg:,}")
        k4.metric("% conv. Leads",     f"{avg_cl:.2f}%",    help="Inscrits / Clics GSC")
        k5.metric("% conv. Bookings",  f"{avg_cb:.2f}%",    help="Bookings / Inscrits")
        k6.metric("💶 AOV moyen",      f"{avg_aov:,.0f} €", help="TTV / Bookings")

        st.divider()

        # ── Scatter — matrice de conversion ──
        if has_ins:
            st.subheader("🔵 Matrice de conversion · % conv. Leads vs % conv. Bookings")
            st.caption("Taille des bulles = TTV · Pages en haut à droite = meilleures des deux dimensions")
            ds = fa[fa["conv_leads_pct"].notna() & fa["conv_bkg_pct"].notna() & (fa["TTV"] > 0)].copy()
            if not ds.empty:
                fig_sc = px.scatter(
                    ds, x="conv_leads_pct", y="conv_bkg_pct", size="TTV",
                    color="type_page" if "type_page" in ds.columns else None,
                    hover_name="url_label",
                    hover_data={"clicks":True,"Inscrits":True,"Bookings":True,
                                "TTV":":.0f","aov":":.0f","conv_leads_pct":":.2f","conv_bkg_pct":":.2f"},
                    labels={"conv_leads_pct":"% conv. Leads (Inscrits/Clics)",
                            "conv_bkg_pct":"% conv. Bookings (Bookings/Inscrits)",
                            "type_page":"Type"},
                    size_max=60, height=600
                )
                mx = ds["conv_leads_pct"].median()
                my = ds["conv_bkg_pct"].median()
                fig_sc.add_vline(x=mx, line_dash="dash", line_color="gray", opacity=0.5,
                                 annotation_text=f"Médiane {mx:.2f}%", annotation_position="top right")
                fig_sc.add_hline(y=my, line_dash="dash", line_color="gray", opacity=0.5,
                                 annotation_text=f"Médiane {my:.2f}%", annotation_position="top right")
                st.plotly_chart(fig_sc, use_container_width=True)

        st.divider()

        # ── Bar chart trié par métrique ──
        st.subheader("📊 Top pages par métrique de conversion")
        fm = st.radio("Trier par :",
                      ["% conv. Leads","% conv. Bookings","AOV (€)","Clics GSC","Inscrits"],
                      horizontal=True, key="fmr")
        col_m = {"% conv. Leads":"conv_leads_pct","% conv. Bookings":"conv_bkg_pct",
                 "AOV (€)":"aov","Clics GSC":"clicks","Inscrits":"Inscrits"}[fm]

        df_bar = fa[fa[col_m].notna()].sort_values(col_m, ascending=False).head(top_n).sort_values(col_m)
        fig_bar = px.bar(df_bar, x=col_m, y="url_label", orientation="h", color=col_m,
                         color_continuous_scale="Blues", title=f"Top {top_n} · {fm}",
                         height=max(400, len(df_bar)*28),
                         hover_data={"clicks":True,"Inscrits":True,"Bookings":True,"TTV":":.0f","aov":":.0f"},
                         labels={col_m: fm, "url_label": "Page"})
        fig_bar.update_layout(coloraxis_showscale=False, yaxis=dict(tickfont=dict(size=10)),
                              margin=dict(l=10,r=40,t=50,b=20))
        st.plotly_chart(fig_bar, use_container_width=True)

        st.divider()

        # ── Vue par type de page ──
        if "type_page" in fa.columns and fa["type_page"].notna().any():
            st.subheader("🗂️ Métriques funnel par type de page")
            fa_t = fa[fa["type_page"] != "Autre"].copy() if st.checkbox("Masquer 'Autre'", value=True, key="haf") else fa.copy()

            tf = fa_t.groupby("type_page").agg(
                Clics=("clicks","sum"), Inscrits=("Inscrits","sum"),
                Bookings=("Bookings","sum"), TTV=("TTV","sum")
            ).reset_index()
            tf["conv_leads_pct"] = (tf["Inscrits"] / tf["Clics"].replace(0,float("nan")) * 100).round(2)
            tf["conv_bkg_pct"]   = (tf["Bookings"] / tf["Inscrits"].replace(0,float("nan")) * 100).round(2)
            tf["aov"]            = (tf["TTV"] / tf["Bookings"].replace(0,float("nan"))).round(0)
            tf = tf.sort_values("TTV", ascending=False)

            fc1,fc2,fc3 = st.columns(3)
            for col, val_c, title, scale in [
                (fc1,"conv_leads_pct","% conv. Leads par type","Blues"),
                (fc2,"conv_bkg_pct","% conv. Bookings par type","Oranges"),
                (fc3,"aov","AOV (€) par type","Greens"),
            ]:
                with col:
                    fig_tf = px.bar(tf.sort_values(val_c), x=val_c, y="type_page", orientation="h",
                                    title=title, color=val_c, color_continuous_scale=scale,
                                    height=max(350,len(tf)*40), labels={val_c:title.split(" par")[0],"type_page":"Type"})
                    fig_tf.update_layout(coloraxis_showscale=False)
                    st.plotly_chart(fig_tf, use_container_width=True)

            st.dataframe(
                tf[["type_page","Clics","Inscrits","Bookings","TTV","conv_leads_pct","conv_bkg_pct","aov"]]
                .rename(columns={"type_page":"Type","conv_leads_pct":"% conv. Leads",
                                  "conv_bkg_pct":"% conv. Bookings","aov":"AOV (€)"}),
                hide_index=True, use_container_width=True
            )

        st.divider()

        # ── Tableau détaillé ──
        st.subheader("📋 Tableau détaillé — Funnel par URL")
        fd_disp = fa.sort_values("TTV", ascending=False).head(top_n).copy()
        fd_disp["TTV (€)"]          = fd_disp["TTV"].map(lambda x: f"{x:,.0f}")
        fd_disp["AOV (€)"]          = fd_disp["aov"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
        fd_disp["% conv. Leads"]    = fd_disp["conv_leads_pct"].map(lambda x: f"{x:.2f}%" if pd.notna(x) else "—")
        fd_disp["% conv. Bookings"] = fd_disp["conv_bkg_pct"].map(lambda x: f"{x:.2f}%" if pd.notna(x) else "—")
        disp_cols = (["url_label","type_page"] if "type_page" in fd_disp.columns else ["url_label"]) + \
                    ["clicks","Inscrits","Bookings","TTV (€)","% conv. Leads","% conv. Bookings","AOV (€)"]
        st.dataframe(fd_disp[disp_cols].rename(columns={"url_label":"Page","type_page":"Type","clicks":"Clics GSC"}),
                     hide_index=True, use_container_width=True)

        csv_f = fa[["vp_url_norm","url_label","type_page","clicks","Inscrits","Bookings","TTV",
                    "conv_leads_pct","conv_bkg_pct","aov"]].rename(columns={
            "vp_url_norm":"URL","url_label":"Label","type_page":"Type","clicks":"Clics GSC",
            "conv_leads_pct":"% conv. Leads","conv_bkg_pct":"% conv. Bookings","aov":"AOV (€)"
        }).to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Exporter CSV Funnel", data=csv_f,
                           file_name="vp_funnel_conversion.csv", mime="text/csv")

# ──────────────────────────────────────────
# TABLEAU RÉCAP GLOBAL + EXPORTS
# ──────────────────────────────────────────

st.divider()
st.subheader("📊 Tableau récap · TTV + Bookings (période sélectionnée)")

df_recap = df_merged_global.sort_values("TTV", ascending=False).copy()
df_recap["TTV / Booking (€)"] = (df_recap["TTV"] / df_recap["Bookings"].replace(0, float("nan"))).round(0)

if "type_page" in df_filtered.columns:
    df_recap = df_recap.merge(
        df_filtered[["vp_url","type_page","destination"]].drop_duplicates("vp_url"),
        on="vp_url", how="left"
    )

dr = df_recap.head(top_n).copy()
dr["TTV (€)"]           = dr["TTV"].map(lambda x: f"{x:,.0f}")
dr["Bookings"]          = dr["Bookings"].map(lambda x: f"{int(x):,}")
dr["TTV / Booking (€)"] = dr["TTV / Booking (€)"].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
dr = dr.rename(columns={"campaign_id":"ID","campaign_name":"Nom campagne","vp_url":"URL VP"})
extra = ["Type","Destination"] if "type_page" in dr.columns else []
if extra:
    dr = dr.rename(columns={"type_page":"Type","destination":"Destination"})
st.caption(f"Top {top_n} pages · TTV décroissant")
st.dataframe(dr[["ID","Nom campagne","URL VP"] + extra + ["TTV (€)","Bookings","TTV / Booking (€)"]],
             hide_index=True, use_container_width=True)

dl1, dl2 = st.columns(2)

with dl1:
    _cc = ["campaign_id","campaign_name","vp_url","TTV","Bookings","TTV / Booking (€)"]
    if "type_page" in df_recap.columns:
        _cc += ["type_page","destination"]
    st.download_button("⬇️ Exporter CSV",
                       data=df_recap[_cc].rename(columns={"type_page":"Schéma","destination":"Destination"}).to_csv(index=False).encode(),
                       file_name="vp_seo_recap.csv", mime="text/csv")

with dl2:
    def build_excel(df_r, df_ttv, df_bkg, df_tpl, month_labels, sel_months):
        wb = Workbook()
        _B, _G, _O = "1d6fa4", "1a7a4a", "f4a840"

        def hdr(ws, color, row=1):
            for cell in ws[row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center")

        def aw(ws, pad=4):
            MAX = {"Nom campagne":40,"URL VP":50,"URL":50,"Page":50}
            for col in ws.columns:
                h = str(col[0].value or "")
                w = min(max(len(str(c.value or "")) for c in col)+pad, MAX.get(h,999))
                ws.column_dimensions[get_column_letter(col[0].column)].width = w

        # Onglet 1 : URLs
        ws1 = wb.active; ws1.title = "URLs"
        hc  = "type_page" in df_r.columns
        heads = ["ID","Nom campagne","URL VP"] + (["Schéma de recherche","Destination"] if hc else []) + ["TTV (€)","Bookings","TTV/Booking (€)"]
        ws1.append(["Période : " + ", ".join([month_labels[m] for m in sorted(sel_months)])])
        ws1[1][0].font = Font(italic=True, color="555555")
        ws1.append(heads); hdr(ws1, _B, row=2)
        for _, row in df_r.iterrows():
            tvb  = row["TTV"]/row["Bookings"] if row["Bookings"] > 0 else None
            base = [int(row["campaign_id"]), row["campaign_name"], row["vp_url"]]
            cats = [row.get("type_page",""), row.get("destination","")] if hc else []
            ws1.append(base + cats + [round(row["TTV"],0), int(row["Bookings"]), round(tvb,0) if tvb else ""])
        aw(ws1)

        # Onglet 2 : Récap Schéma
        if hc:
            ws2 = wb.create_sheet("Récap Schéma")
            ws2.append(["Schéma","TTV (€)","Bookings","TTV/Booking (€)","TTV/Lead (€)"]); hdr(ws2, _B)
            _st = df_r.groupby("type_page")["TTV"].sum()
            _sb = df_r.groupby("type_page")["Bookings"].sum()
            _tpl_s = {}
            if not df_tpl.empty and "type_page" in df_filtered.columns:
                _tm = df_tpl.merge(df_filtered[["campaign_id","type_page"]].drop_duplicates(), on="campaign_id", how="left")
                _tpl_s = _tm.groupby("type_page")["value"].mean().to_dict()
            for s in _st.index:
                t = round(_st[s],0); b = int(_sb.get(s,0))
                ws2.append([s, t, b, round(t/b,0) if b>0 else "", round(_tpl_s.get(s,0),2) if s in _tpl_s else ""])
            aw(ws2)

            # Onglet 3 : Récap Destination
            ws3 = wb.create_sheet("Récap Destination")
            ws3.append(["Destination","TTV (€)","Bookings","TTV/Booking (€)","TTV/Lead (€)"]); hdr(ws3, _G)
            _dt = df_r.groupby("destination")["TTV"].sum()
            _db = df_r.groupby("destination")["Bookings"].sum()
            _tpl_d = {}
            if not df_tpl.empty and "destination" in df_filtered.columns:
                _dm = df_tpl.merge(df_filtered[["campaign_id","destination"]].drop_duplicates(), on="campaign_id", how="left")
                _tpl_d = _dm.groupby("destination")["value"].mean().to_dict()
            for d in _dt.index:
                t = round(_dt[d],0); b = int(_db.get(d,0))
                ws3.append([d, t, b, round(t/b,0) if b>0 else "", round(_tpl_d.get(d,0),2) if d in _tpl_d else ""])
            aw(ws3)

        # Onglet 4 : TTV par mois
        ws4 = wb.create_sheet("TTV par mois")
        mnums = sorted(set(df_ttv["month"].tolist()))
        ws4.append(["URL"] + [month_labels[m] for m in mnums] + ["TOTAL"]); hdr(ws4, _B)
        piv4 = df_ttv.groupby(["vp_url","month"])["value"].sum().reset_index()\
                     .pivot(index="vp_url",columns="month",values="value").fillna(0)
        piv4["TOTAL"] = piv4.sum(axis=1)
        piv4 = piv4.sort_values("TOTAL", ascending=False)
        for url, row in piv4.iterrows():
            ws4.append([url] + [round(row.get(m,0),0) for m in mnums] + [round(row["TOTAL"],0)])
        aw(ws4)

        # Onglet 5 : Bookings par mois
        ws5 = wb.create_sheet("Bookings par mois")
        ws5.append(["URL"] + [month_labels[m] for m in mnums] + ["TOTAL"]); hdr(ws5, _O)
        piv5 = df_bkg.groupby(["vp_url","month"])["value"].sum().reset_index()\
                     .pivot(index="vp_url",columns="month",values="value").fillna(0)
        piv5["TOTAL"] = piv5.sum(axis=1)
        piv5 = piv5.sort_values("TOTAL", ascending=False)
        for url, row in piv5.iterrows():
            ws5.append([url] + [int(row.get(m,0)) for m in mnums] + [int(row["TOTAL"])])
        aw(ws5)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf.read()

    xls = build_excel(df_recap, df_ttv_agg, df_bkg_agg, df_tpl_agg, MONTH_LABELS, selected_months)
    st.download_button(
        "⬇️ Exporter Excel (5 onglets)", data=xls,
        file_name=f"vp_seo_{'_'.join([MONTH_LABELS[m][:3] for m in sorted(selected_months)])}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
